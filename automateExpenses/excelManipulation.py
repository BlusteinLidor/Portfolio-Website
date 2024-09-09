from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# load the excel workbook
wb = load_workbook("transaction-details_export_1725905023405.xlsx")
# get the active (default) worksheet
ws = wb.active

categoryCol = 'C'
expenseCol = 'F'
categoryRow = 5
categoryDict = dict()
while(ws[categoryCol + str(categoryRow)].value is not None):
    if ws[categoryCol + str(categoryRow)].value in categoryDict:
        categoryDict[ws[categoryCol + str(categoryRow)].value] += ws[expenseCol + str(categoryRow)].value
    else:
        categoryDict.update({ws[categoryCol + str(categoryRow)].value : ws[expenseCol + str(categoryRow)].value})
    categoryRow += 1

i = 1
print(categoryDict)
print("the length of the set is:" , str(len(categoryDict)) , "\nand the set items are:")
for category in categoryDict:
    print(i,".",category)
    i += 1

def getLargestExpense():
    largetExpenseCategory = max(zip(categoryDict.values(), categoryDict.keys()))[1]  
    largetExpenseAmount = categoryDict[largetExpenseCategory]
    print("The category with the largest expense is :", largetExpenseCategory, "with", largetExpenseAmount, "Shekels")

getLargestExpense()